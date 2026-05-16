#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从 IbatisToJdbcConverterTest 自动提取：
1) 场景：@Test 上方注释第一行
2) 入参：converter.convertPrepared 的第二个参数
3) 原始 XML：按 statementId 从 test-sqlmap.xml 提取对应节点
4) 得到的 SQL：优先取 assertEquals("...", var.getSql())，否则标记为 N/A
"""

import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path("/Users/xusheng/Desktop/blackbox/ibatis2jdbc")
TEST_JAVA = ROOT / "src/test/java/com/blackbox/ibatis2jdbc/IbatisToJdbcConverterTest.java"
SQLMAP_XML = ROOT / "src/test/resources/sqlmaps/test-sqlmap.xml"
OUTPUT = ROOT / "IbatisToJdbcConverter_场景测试.xlsx"


def split_top_level_args(arg_text: str):
    args = []
    current = []
    depth_paren = 0
    depth_bracket = 0
    depth_brace = 0
    in_single = False
    in_double = False
    escape = False

    for ch in arg_text:
        if escape:
            current.append(ch)
            escape = False
            continue

        if ch == "\\":
            current.append(ch)
            escape = True
            continue

        if in_single:
            current.append(ch)
            if ch == "'":
                in_single = False
            continue

        if in_double:
            current.append(ch)
            if ch == '"':
                in_double = False
            continue

        if ch == "'":
            in_single = True
            current.append(ch)
            continue

        if ch == '"':
            in_double = True
            current.append(ch)
            continue

        if ch == "(":
            depth_paren += 1
            current.append(ch)
            continue
        if ch == ")":
            depth_paren -= 1
            current.append(ch)
            continue
        if ch == "[":
            depth_bracket += 1
            current.append(ch)
            continue
        if ch == "]":
            depth_bracket -= 1
            current.append(ch)
            continue
        if ch == "{":
            depth_brace += 1
            current.append(ch)
            continue
        if ch == "}":
            depth_brace -= 1
            current.append(ch)
            continue

        if ch == "," and depth_paren == 0 and depth_bracket == 0 and depth_brace == 0:
            args.append("".join(current).strip())
            current = []
            continue

        current.append(ch)

    tail = "".join(current).strip()
    if tail:
        args.append(tail)
    return args


def find_matching_paren(text: str, open_index: int):
    depth = 0
    in_single = False
    in_double = False
    escape = False

    for i in range(open_index, len(text)):
        ch = text[i]

        if escape:
            escape = False
            continue

        if ch == "\\":
            escape = True
            continue

        if in_single:
            if ch == "'":
                in_single = False
            continue

        if in_double:
            if ch == '"':
                in_double = False
            continue

        if ch == "'":
            in_single = True
            continue
        if ch == '"':
            in_double = True
            continue

        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
            if depth == 0:
                return i

    return -1


def normalize_scene_line(line: str):
    line = line.strip()
    line = re.sub(r"^\*+", "", line).strip()
    return line


def extract_first_javadoc_line(lines, at_index):
    i = at_index - 1
    while i >= 0 and lines[i].strip() == "":
        i -= 1
    if i < 0:
        return ""

    if "*/" not in lines[i]:
        return ""

    end = i
    start = -1
    while i >= 0:
        if "/**" in lines[i]:
            start = i
            break
        i -= 1

    if start == -1:
        return ""

    for j in range(start + 1, end + 1):
        candidate = normalize_scene_line(lines[j])
        if candidate and candidate != "/" and not candidate.startswith("@"):
            return candidate
    return ""


def extract_method_name(lines, at_index):
    for i in range(at_index + 1, min(len(lines), at_index + 10)):
        line = lines[i].strip()
        m = re.search(r"(?:public\s+)?void\s+(\w+)\s*\(", line)
        if m:
            return m.group(1), i
    return "", -1


def extract_method_block(lines, method_line_index):
    joined = "\n".join(lines)
    pos = 0
    for i in range(method_line_index):
        pos += len(lines[i]) + 1
    method_text = joined[pos:]

    open_brace = method_text.find("{")
    if open_brace == -1:
        return ""

    abs_open = pos + open_brace
    depth = 0
    in_single = False
    in_double = False
    escape = False

    for idx in range(abs_open, len(joined)):
        ch = joined[idx]

        if escape:
            escape = False
            continue

        if ch == "\\":
            escape = True
            continue

        if in_single:
            if ch == "'":
                in_single = False
            continue

        if in_double:
            if ch == '"':
                in_double = False
            continue

        if ch == "'":
            in_single = True
            continue
        if ch == '"':
            in_double = True
            continue

        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                return joined[abs_open + 1:idx]

    return ""


def extract_sql_assert_map(method_body: str):
    pattern = re.compile(
        r'assertEquals\(\s*"((?:\\.|[^"\\])*)"\s*,\s*(\w+)\.getSql\(\)\s*\);',
        re.S,
    )
    result = {}
    for m in pattern.finditer(method_body):
        sql = m.group(1).replace('\\"', '"')
        var = m.group(2)
        result[var] = sql
    return result


def extract_comment_above_index(text: str, index: int):
    before = text[:index].rstrip()
    if not before:
        return ""

    lines = before.splitlines()
    i = len(lines) - 1

    while i >= 0 and lines[i].strip() == "":
        i -= 1

    if i < 0:
        return ""

    line = lines[i].strip()
    if line.startswith("//"):
        return line[2:].strip()
    if line.startswith("/*"):
        line = line[2:].strip()
        if line.endswith("*/"):
            line = line[:-2].strip()
        return normalize_scene_line(line)
    if line.startswith("*"):
        return normalize_scene_line(line)
    return ""


def extract_convert_calls(method_body: str):
    calls = []
    marker = "converter.convertPrepared("
    start = 0

    while True:
        idx = method_body.find(marker, start)
        if idx == -1:
            break

        open_idx = idx + len("converter.convertPrepared")
        close_idx = find_matching_paren(method_body, open_idx)
        if close_idx == -1:
            break

        call_text = method_body[idx:close_idx + 1]
        args_text = method_body[open_idx + 1:close_idx]
        args = split_top_level_args(args_text)

        line_start = method_body.rfind("\n", 0, idx)
        if line_start == -1:
            line_start = 0
        prev_stmt_start = method_body.rfind(";", 0, idx)
        segment_start = max(line_start, prev_stmt_start + 1)
        prefix = method_body[segment_start:idx]

        m_var = re.search(r"ConvertedSql\s+(\w+)\s*=\s*$", prefix.strip())
        var_name = m_var.group(1) if m_var else ""
        decl_match = re.search(r"ConvertedSql\s+\w+\s*=\s*$", prefix, re.S)
        decl_start = segment_start + decl_match.start() if decl_match else idx

        calls.append({
            "args": args,
            "var": var_name,
            "call_text": call_text,
            "start": idx,
            "end": close_idx + 1,
            "decl_start": decl_start,
        })

        start = close_idx + 1

    return calls


def strip_java_string_quotes(text: str):
    text = text.strip()
    if len(text) >= 2 and text[0] == '"' and text[-1] == '"':
        return text[1:-1]
    return text


def extract_statement_xml_map(xml_text: str):
    statement_xml = {}
    pattern = re.compile(
        r"<(select|update|delete)\b[^>]*\bid\s*=\s*\"([^\"]+)\"[^>]*>.*?</\1>",
        re.S,
    )
    for m in pattern.finditer(xml_text):
        statement_id = m.group(2)
        raw = m.group(0).strip()
        statement_xml[statement_id] = raw
    return statement_xml


def build_rows():
    java_text = TEST_JAVA.read_text(encoding="utf-8")
    xml_text = SQLMAP_XML.read_text(encoding="utf-8")
    statement_xml_map = extract_statement_xml_map(xml_text)

    lines = java_text.splitlines()
    rows = []
    seq = 1

    for i, line in enumerate(lines):
        if line.strip() != "@Test":
            continue

        scene = extract_first_javadoc_line(lines, i)
        method_name, method_line_idx = extract_method_name(lines, i)
        if method_line_idx == -1:
            continue

        method_body = extract_method_block(lines, method_line_idx)
        calls = extract_convert_calls(method_body)

        call_idx = 0
        for call_index, call in enumerate(calls):
            args = call["args"]
            var_name = call["var"]
            start_index = call["start"]
            end_index = call["end"]
            decl_start = call["decl_start"]

            if len(args) == 2:
                statement_raw = args[0]
                input_param = args[1].strip()
            else:
                statement_raw = ""
                input_param = ""

            statement_id = strip_java_string_quotes(statement_raw)
            original_xml = statement_xml_map.get(statement_id, "N/A")

            next_start = calls[call_index + 1]["start"] if call_index + 1 < len(calls) else len(method_body)
            search_block = method_body[end_index:next_start]
            # 优先查 toPreviewSql()，再查 getSql()
            sql_match = re.search(
                rf'assertEquals\(\s*"((?:\\.|[^"\\])*)"\s*,\s*{re.escape(var_name)}\.toPreviewSql\(\)\s*\);',
                search_block,
                re.S,
            )
            if not sql_match:
                sql_match = re.search(
                    rf'assertEquals\(\s*"((?:\\.|[^"\\])*)"\s*,\s*{re.escape(var_name)}\.getSql\(\)\s*\);',
                    search_block,
                    re.S,
                )
            sql = sql_match.group(1).replace('\\"', '"') if sql_match else "N/A"

            sub_scene = extract_comment_above_index(method_body, decl_start)
            if not sub_scene:
                sub_scene = ""

            call_idx += 1

            rows.append({
                "序号": seq,
                "场景": scene if scene else method_name,
                "子场景": sub_scene,
                "入参": input_param,
                "原始 XML": original_xml,
                "得到的 SQL": sql,
            })
            seq += 1

    return rows


def create_excel(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IbatisToJdbcConverterTest"

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 44
    ws.column_dimensions["E"].width = 78
    ws.column_dimensions["F"].width = 78

    ws.row_dimensions[1].height = 30

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    headers = ["序号", "场景", "子场景", "入参", "原始 XML", "得到的 SQL"]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    for row_num, case in enumerate(rows, 2):
        ws.row_dimensions[row_num].height = 48
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = case[header]
            cell.alignment = cell_alignment
            cell.border = border

    ws.freeze_panes = "A2"
    wb.save(OUTPUT)


def main():
    rows = build_rows()
    create_excel(rows)
    print(f"✅ Excel 文件已生成：{OUTPUT}")
    print(f"📊 包含 {len(rows)} 个转换结果")


if __name__ == "__main__":
    main()
