#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将 Markdown 测试报告导出为 Excel。

默认输入：sqlmap-classpath-scan-full-report.md
默认输出：sqlmap-classpath-scan-full-report.xlsx

依赖：
- openpyxl
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


SKIP_SECTION_TITLES = {"测试时间", "机器参数", "扫描与测试统计", "扫描阶段性能", "按 XML 维度转换统计"}


def extract_statement_xml_value(statement_id: str, xml_block: str) -> str:
    """从整份 sqlMap XML 中提取指定 statement 的完整 XML（含外层标签）。"""
    if not xml_block:
        return ""

    local_id = statement_id.rsplit(".", 1)[-1] if statement_id else ""
    candidate_ids = [statement_id, local_id]

    for candidate in candidate_ids:
        if not candidate:
            continue
        pattern = re.compile(
            r"<(?P<tag>select|insert|update|delete|statement)\b[^>]*\bid\s*=\s*\""
            + re.escape(candidate)
            + r"\"[^>]*>(?P<body>[\s\S]*?)</(?P=tag)>",
            re.IGNORECASE,
        )
        match = pattern.search(xml_block)
        if match:
            return match.group(0).strip()

    return xml_block.strip()


def parse_sections(markdown: str) -> Dict[str, List[str]]:
    """按 markdown 二级标题分段。"""
    sections: Dict[str, List[str]] = {}
    current = "_top"
    sections[current] = []
    for line in markdown.splitlines():
        if line.startswith("## "):
            current = line[3:].strip()
            sections.setdefault(current, [])
            continue
        sections[current].append(line)
    return sections


def parse_bullet_kv(lines: List[str]) -> List[Tuple[str, str]]:
    """解析 '- key：value' 形式的行。"""
    out: List[Tuple[str, str]] = []
    for line in lines:
        text = line.strip()
        if not text.startswith("- "):
            continue
        body = text[2:].strip()
        if "：" in body:
            k, v = body.split("：", 1)
        elif ":" in body:
            k, v = body.split(":", 1)
        else:
            k, v = body, ""
        out.append((k.strip(), v.strip()))
    return out


def parse_markdown_table(lines: List[str]) -> List[List[str]]:
    """解析 markdown 表格（简单场景）。"""
    table_rows: List[List[str]] = []
    in_table = False
    for line in lines:
        if not line.strip().startswith("|"):
            if in_table:
                break
            continue
        in_table = True
        cols = [c.strip() for c in line.strip().strip("|").split("|")]
        # 跳过分隔行
        if all(re.fullmatch(r"-+:?", c.replace(" ", "")) for c in cols):
            continue
        table_rows.append(cols)
    return table_rows


def split_by_h2(markdown: str) -> List[Tuple[str, List[str]]]:
    """按二级标题切分全文，返回 [(title, lines)]。"""
    parts: List[Tuple[str, List[str]]] = []
    current_title = "_top"
    current_lines: List[str] = []
    for line in markdown.splitlines():
        if line.startswith("## "):
            parts.append((current_title, current_lines))
            current_title = line[3:].strip()
            current_lines = []
            continue
        current_lines.append(line)
    parts.append((current_title, current_lines))
    return parts


def extract_fenced_block(lines: List[str], start_index: int) -> Tuple[str, int]:
    """从 start_index 起提取 ``` fenced block 内容。"""
    i = start_index
    while i < len(lines) and not lines[i].strip().startswith("```"):
        i += 1
    if i >= len(lines):
        return "", start_index

    i += 1
    buf: List[str] = []
    while i < len(lines):
        if lines[i].strip().startswith("```"):
            return "\n".join(buf).strip(), i
        buf.append(lines[i])
        i += 1
    return "\n".join(buf).strip(), i


def parse_statement_blocks(markdown: str) -> List[Dict[str, str]]:
    """解析每个 statement 的详细报告块（含入参/XML/SQL/错误）。"""
    blocks: List[Dict[str, str]] = []
    for title, lines in split_by_h2(markdown):
        if title in SKIP_SECTION_TITLES or title == "_top":
            continue

        item: Dict[str, str] = {
            "statementId": title,
            "source": "",
            "statementType": "",
            "elapsedMs": "",
            "result": "",
            "parameters": "",
            "originalXml": "",
            "finalSql": "",
            "error": "",
        }

        i = 0
        while i < len(lines):
            row = lines[i].strip()
            if row.startswith("- source："):
                item["source"] = row.split("：", 1)[1].strip()
            elif row.startswith("- statement 类型："):
                item["statementType"] = row.split("：", 1)[1].strip()
            elif row.startswith("- 转换耗时(ms)："):
                item["elapsedMs"] = row.split("：", 1)[1].strip()
            elif row.startswith("- 结果："):
                item["result"] = row.split("：", 1)[1].strip()
            elif row.startswith("- 入参："):
                block, end_idx = extract_fenced_block(lines, i + 1)
                item["parameters"] = block
                i = end_idx
            elif row.startswith("- 原始 XML："):
                block, end_idx = extract_fenced_block(lines, i + 1)
                item["originalXml"] = block
                i = end_idx
            elif row.startswith("- 最终 SQL"):
                block, end_idx = extract_fenced_block(lines, i + 1)
                item["finalSql"] = block
                i = end_idx
            elif row.startswith("- 错误："):
                block, end_idx = extract_fenced_block(lines, i + 1)
                item["error"] = block
                i = end_idx
            i += 1

        if item["statementId"]:
            blocks.append(item)
    return blocks


def export_excel(md_path: Path, xlsx_path: Path) -> None:
    text = md_path.read_text(encoding="utf-8")
    sections = parse_sections(text)

    wb = Workbook()

    # 概览页
    ws = wb.active
    ws.title = "概览"
    ws.append(["分类", "指标", "值"])

    for sec_name in ["测试时间", "机器参数", "扫描与测试统计", "扫描阶段性能"]:
        for k, v in parse_bullet_kv(sections.get(sec_name, [])):
            ws.append([sec_name, k, v])

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 80
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # XML 统计页
    ws2 = wb.create_sheet("XML统计")
    table_rows = parse_markdown_table(sections.get("按 XML 维度转换统计", []))
    if table_rows:
        for r in table_rows:
            ws2.append(r)
    else:
        ws2.append(["无表格数据"])

    for col in ["A", "B", "C", "D", "E", "F"]:
        ws2.column_dimensions[col].width = 28
    if ws2.max_row >= 1:
        for cell in ws2[1]:
            cell.font = Font(bold=True)
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Statement 明细页（包含入参/XML/SQL/错误）
    ws3 = wb.create_sheet("Statement明细")
    data_keys = [
        "statementId",
        "source",
        "statementType",
        "elapsedMs",
        "result",
        "parameters",
        "originalXml",
        "finalSql",
        "error",
    ]
    headers = [
        "语句ID",
        "来源",
        "语句类型",
        "耗时(ms)",
        "结果",
        "入参",
        "原始XML",
        "最终SQL",
        "错误",
    ]
    ws3.append(headers)
    for block in parse_statement_blocks(text):
        ws3.append([block.get(k, "") for k in data_keys])

    ws3.column_dimensions["A"].width = 40
    ws3.column_dimensions["B"].width = 32
    ws3.column_dimensions["C"].width = 16
    ws3.column_dimensions["D"].width = 12
    ws3.column_dimensions["E"].width = 10
    ws3.column_dimensions["F"].width = 40
    ws3.column_dimensions["G"].width = 50
    ws3.column_dimensions["H"].width = 50
    ws3.column_dimensions["I"].width = 40
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    for row in ws3.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(xlsx_path)




def build_paths(input_md: Path, out_xlsx: str | None) -> Path:
    stem = input_md.with_suffix("")
    xlsx_path = Path(out_xlsx) if out_xlsx else Path(str(stem) + ".xlsx")
    return xlsx_path


def main() -> None:
    parser = argparse.ArgumentParser(description="将测试 markdown 报告导出为 Excel")
    parser.add_argument("--input", default="sqlmap-classpath-scan-full-report.md", help="输入 markdown 文件")
    parser.add_argument("--excel", default=None, help="输出 xlsx 文件路径")
    args = parser.parse_args()

    md_path = Path(args.input)
    if not md_path.exists():
        raise FileNotFoundError(f"未找到输入文件: {md_path}")

    xlsx_path = build_paths(md_path, args.excel)
    export_excel(md_path, xlsx_path)

    print(f"Excel 已生成: {xlsx_path}")


if __name__ == "__main__":
    main()
